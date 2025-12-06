from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import cache_control
from django.http import JsonResponse
from django.utils import timezone
from django.db import models
from .models import DailyRecordSIAF, FemaleBirdsMortality,FemaleBirdsStock, FeedStock, MaleBirdsStock, MaleBirdsMortality
from datetime import datetime, timedelta
import pandas as pd
from django.contrib.auth.models import User, Group
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("dashboard")   # ✅ always redirect to dashboard
        else:
            messages.error(request, "Invalid username or password")

    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def dashboard(request):

    user_groups = request.user.groups.all()
    u = request.user

    try:
        # Get selected date or default to today
        selected_date = request.GET.get('date')
        if selected_date:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        else:
            selected_date = timezone.now().date()
    except ValueError:
        selected_date = timezone.now().date()
    
    # Get SIAF (SIAF) record
    siaf_record = DailyRecordSIAF.objects.filter(date=selected_date).first()
    
    # Calculate totals for SIAF
    siaf_total_tray_eggs = 0
    siaf_total_eggs = 0
    if siaf_record:
        siaf_total_tray_eggs = (siaf_record.tray_egg_morning or 0) + (siaf_record.tray_egg_evening or 0)
        siaf_total_eggs = (siaf_record.total_egg_morning or 0) + (siaf_record.total_egg_evening or 0)
    
    # Set default values for bird counts (no BirdsCount model - removed)
    initial_siaf_birds = 0
    current_siaf_birds = 0
    siaf_mortality = 0
    
    # Calculate percentages
    eggs_per_current_birds = 0
    eggs_per_initial_birds = 0
    
    # Calculate male and female bird counts for active batches
    male_current_birds = 0
    female_current_birds = 0
    total_current_birds = 0
    male_total_mortality = 0
    female_total_mortality = 0
    total_mortality = 0
    
    active_male_batches = MaleBirdsStock.objects.filter(status='active')
    active_female_batches = FemaleBirdsStock.objects.filter(status='active')
    
    for batch in active_male_batches:
        male_current_birds += batch.get_current_birds()
        male_total_mortality += batch.get_current_mortality()
    
    for batch in active_female_batches:
        female_current_birds += batch.get_current_birds()
        female_total_mortality += batch.get_current_mortality()
    
    total_current_birds = male_current_birds + female_current_birds
    total_mortality = male_total_mortality + female_total_mortality
    
    # Calculate feed per gram per bird
    feed_per_gram_per_bird = 0
    if siaf_record and total_current_birds > 0:
        total_feed_kg = (siaf_record.feed_morning or 0) + (siaf_record.feed_evening or 0)
        total_feed_grams = total_feed_kg * 1000
        feed_per_gram_per_bird = round(total_feed_grams / total_current_birds, 2)
    
    context = {
        'user_groups': user_groups,
        'u': u,
        'today_date': selected_date,
        'siaf_total_tray_eggs': siaf_total_tray_eggs,
        'siaf_total_eggs': siaf_total_eggs,
        'siaf_record_exists': siaf_record is not None,
        'current_siaf_birds': current_siaf_birds,
        'initial_siaf_birds': initial_siaf_birds,
        'eggs_per_current_birds': round(eggs_per_current_birds, 2),
        'eggs_per_initial_birds': round(eggs_per_initial_birds, 2),
        'male_current_birds': male_current_birds,
        'female_current_birds': female_current_birds,
        'total_current_birds': total_current_birds,
        'male_total_mortality': male_total_mortality,
        'female_total_mortality': female_total_mortality,
        'total_mortality': total_mortality,
        'feed_per_gram_per_bird': feed_per_gram_per_bird
    }
    
    return render(request, "dashboard.html", context)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def report(request):
    user_groups = request.user.groups.all()
    u = request.user
    return render(request, "report.html", {'user_groups': user_groups, 'u': u})

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def feed(request):
    user_groups = request.user.groups.all()
    u = request.user
    return render(request, "feed.html", {'user_groups': user_groups, 'u': u})

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def males(request):
    user_groups = request.user.groups.all()
    u = request.user
    return render(request, "males.html", {'user_groups': user_groups, 'u': u})

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def females(request):
    user_groups = request.user.groups.all()
    u = request.user
    return render(request, "females.html", {'user_groups': user_groups, 'u': u})

def report_data(request):
    
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Missing required parameters'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Query SIAF (SIAF breed) records for the date range
            records = DailyRecordSIAF.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('-date')

            # Convert records to list of dictionaries
            data = []
            for record in records:
                data.append({
                    'date': record.date,
                    'feed_morning': record.feed_morning,
                    'feed_evening': record.feed_evening,
                    'water_intake': record.water_intake,
                    'tray_egg_morning': record.tray_egg_morning,
                    'tray_egg_evening': record.tray_egg_evening,
                    'total_egg_morning': record.total_egg_morning,
                    'total_egg_evening': record.total_egg_evening,
                })

            return JsonResponse({
                'success': True,
                'records': data
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


# @cache_control(no_cache=True, must_revalidate=True, no_store=True)
# @login_required()
# def SIAF(request):
#     return render(request, "SIAF.html")



@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def SIAF(request):
    user_groups = request.user.groups.all()
    u = request.user

    if request.method == 'POST':
        try:
            # Get or create a record for the specified date
            date_str = request.POST.get('date')
            date = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else timezone.now().date()
            record, created = DailyRecordSIAF.objects.get_or_create(date=date)
            # Feed Data
            record.feed_morning = int(request.POST.get('feed_morning')) if request.POST.get('feed_morning') else None
            record.feed_morning_bundles = float(request.POST.get('feed_morning_bundles')) if request.POST.get('feed_morning_bundles') else None
            record.feed_evening = int(request.POST.get('feed_evening')) if request.POST.get('feed_evening') else None
            record.feed_evening_bundles = float(request.POST.get('feed_evening_bundles')) if request.POST.get('feed_evening_bundles') else None
            record.water_intake = float(request.POST.get('water_intake')) if request.POST.get('water_intake') else None
            
            # Egg Collection Data - Morning
            record.tray_egg_morning = int(request.POST.get('tray_egg_morning')) if request.POST.get('tray_egg_morning') else None
            record.total_egg_morning = int(request.POST.get('total_egg_morning')) if request.POST.get('total_egg_morning') else None
            record.damaged_egg_morning = int(request.POST.get('damaged_egg_morning')) if request.POST.get('damaged_egg_morning') else None
            record.double_egg_morning = int(request.POST.get('double_egg_morning')) if request.POST.get('double_egg_morning') else None
            
            # Egg Collection Data - Evening
            record.tray_egg_evening = int(request.POST.get('tray_egg_evening')) if request.POST.get('tray_egg_evening') else None
            record.total_egg_evening = int(request.POST.get('total_egg_evening')) if request.POST.get('total_egg_evening') else None
            record.damaged_egg_evening = int(request.POST.get('damaged_egg_evening')) if request.POST.get('damaged_egg_evening') else None
            record.double_egg_evening = int(request.POST.get('double_egg_evening')) if request.POST.get('double_egg_evening') else None
            
            # Equipment Status
            record.artificial_insemination = request.POST.get('artificial_insemination', 'No')
            record.ai_hours = float(request.POST.get('ai_hours')) if request.POST.get('ai_hours') else None
            record.fogger_used = request.POST.get('fogger_used', 'No')
            record.fogger_hours = float(request.POST.get('fogger_hours')) if request.POST.get('fogger_hours') else None
            record.fan_used = request.POST.get('fan_used', 'No')
            record.fan_hours = float(request.POST.get('fan_hours')) if request.POST.get('fan_hours') else None
            record.light_used = request.POST.get('light_used', 'No')
            record.light_hours = float(request.POST.get('light_hours')) if request.POST.get('light_hours') else None
            
            # Other Metrics
            record.medicine = request.POST.get('medicine')
            record.notes = request.POST.get('notes')
            
            # Temperature Data (Fahrenheit)
            record.temperature_1 = float(request.POST.get('temperature_1')) if request.POST.get('temperature_1') else None
            record.temperature_2 = float(request.POST.get('temperature_2')) if request.POST.get('temperature_2') else None
            record.temperature_3 = float(request.POST.get('temperature_3')) if request.POST.get('temperature_3') else None
            record.temperature_4 = float(request.POST.get('temperature_4')) if request.POST.get('temperature_4') else None
            record.temperature_5 = float(request.POST.get('temperature_5')) if request.POST.get('temperature_5') else None
            record.temperature_6 = float(request.POST.get('temperature_6')) if request.POST.get('temperature_6') else None
            
            # Save the record
            record.save()
            messages.success(request, 'Daily record for SIAF saved successfully!')
            return redirect('SIAF')  # Stay on the same page
            
        except Exception as e:
            messages.error(request, f'Error saving record: {str(e)}')

    return render(request, "SIAF.html", {'user_groups': user_groups, 'u': u})

@login_required
def dashboard_data(request):
    if request.method == 'GET':
        try:
            date_str = request.GET.get('date')
            if not date_str:
                return JsonResponse({
                    'success': False,
                    'message': 'Date parameter is required'
                })

            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            siaf_record = DailyRecordSIAF.objects.filter(date=date).first()

            # Calculate totals for SIAF
            siaf_total_tray_eggs = 0
            siaf_total_eggs = 0
            if siaf_record:
                siaf_total_tray_eggs = (siaf_record.tray_egg_morning or 0) + (siaf_record.tray_egg_evening or 0)
                siaf_total_eggs = (siaf_record.total_egg_morning or 0) + (siaf_record.total_egg_evening or 0)

            # Calculate male and female bird counts for active batches
            male_current_birds = 0
            female_current_birds = 0
            male_total_mortality = 0
            female_total_mortality = 0
            
            active_male_batches = MaleBirdsStock.objects.filter(status='active')
            active_female_batches = FemaleBirdsStock.objects.filter(status='active')
            
            for batch in active_male_batches:
                male_current_birds += batch.get_current_birds()
                male_total_mortality += batch.get_current_mortality()
            
            for batch in active_female_batches:
                female_current_birds += batch.get_current_birds()
                female_total_mortality += batch.get_current_mortality()
            
            total_current_birds = male_current_birds + female_current_birds
            total_mortality = male_total_mortality + female_total_mortality

            # Calculate feed per gram per bird
            feed_per_gram_per_bird = 0
            if siaf_record and total_current_birds > 0:
                total_feed_kg = (siaf_record.feed_morning or 0) + (siaf_record.feed_evening or 0)
                total_feed_grams = total_feed_kg * 1000
                feed_per_gram_per_bird = round(total_feed_grams / total_current_birds, 2)

            return JsonResponse({
                'success': True,
                'data': {
                    'siaf_total_tray_eggs': siaf_total_tray_eggs,
                    'siaf_total_eggs': siaf_total_eggs,
                    'siaf_record_exists': siaf_record is not None,
                    'male_current_birds': male_current_birds,
                    'female_current_birds': female_current_birds,
                    'total_current_birds': total_current_birds,
                    'male_total_mortality': male_total_mortality,
                    'female_total_mortality': female_total_mortality,
                    'total_mortality': total_mortality,
                    'feed_per_gram_per_bird': feed_per_gram_per_bird
                }
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
@login_required
def download_excel(request):
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not start_date or not end_date:
                return JsonResponse({'success': False, 'message': 'Please provide start and end dates'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Create Excel workbook
            wb = Workbook()
            
            # Define header style
            header_style = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Column headers - Enhanced with temperature, feed per bird, and mortality details
            headers = ['Date', 'Feed Morning', 'Feed Evening', 'Water Intake', 
                      'Tray Egg Morning', 'Total Egg Morning', 'Damaged Egg Morning', 'Double Egg Morning',
                      'Tray Egg Evening', 'Total Egg Evening', 'Damaged Egg Evening', 'Double Egg Evening',
                      'AI Status', 'AI Hours', 'Fogger Status', 'Fogger Hours',
                      'Fan Status', 'Fan Hours','Light Status', 'Light Hours', 
                      'Temp 1 (6AM)', 'Temp 2 (10AM)', 'Temp 3 (2PM)', 'Temp 4 (6PM)', 'Temp 5 (10PM)', 'Temp 6 (2AM)',
                      'Feed Per Bird (g)', 'Male Mortality', 'Female Mortality', 'Total Mortality',
                      'Medicine', 'Notes']

            # Create SIAF sheet
            siaf_sheet = wb.active
            siaf_sheet.title = 'SIAF'
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = siaf_sheet.cell(row=1, column=col, value=header)
                cell.font = header_style
                cell.fill = header_fill

            # Get SIAF (SIAF breed) records
            siaf_records = DailyRecordSIAF.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('date')

            # Add data
            for row, record in enumerate(siaf_records, 2):
                siaf_sheet.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d'))
                siaf_sheet.cell(row=row, column=2, value=record.feed_morning)
                siaf_sheet.cell(row=row, column=3, value=record.feed_evening)
                siaf_sheet.cell(row=row, column=4, value=record.water_intake)
                siaf_sheet.cell(row=row, column=5, value=record.tray_egg_morning)
                siaf_sheet.cell(row=row, column=6, value=record.total_egg_morning)
                siaf_sheet.cell(row=row, column=7, value=record.damaged_egg_morning)
                siaf_sheet.cell(row=row, column=8, value=record.double_egg_morning)
                siaf_sheet.cell(row=row, column=9, value=record.tray_egg_evening)
                siaf_sheet.cell(row=row, column=10, value=record.total_egg_evening)
                siaf_sheet.cell(row=row, column=11, value=record.damaged_egg_evening)
                siaf_sheet.cell(row=row, column=12, value=record.double_egg_evening)
                siaf_sheet.cell(row=row, column=13, value=record.artificial_insemination)
                siaf_sheet.cell(row=row, column=14, value=record.ai_hours)
                siaf_sheet.cell(row=row, column=15, value=record.fogger_used)
                siaf_sheet.cell(row=row, column=16, value=record.fogger_hours)
                siaf_sheet.cell(row=row, column=17, value=record.fan_used)
                siaf_sheet.cell(row=row, column=18, value=record.fan_hours)
                siaf_sheet.cell(row=row, column=19, value=record.light_used)
                siaf_sheet.cell(row=row, column=20, value=record.light_hours)
                # Temperature readings
                siaf_sheet.cell(row=row, column=21, value=record.temperature_1)
                siaf_sheet.cell(row=row, column=22, value=record.temperature_2)
                siaf_sheet.cell(row=row, column=23, value=record.temperature_3)
                siaf_sheet.cell(row=row, column=24, value=record.temperature_4)
                siaf_sheet.cell(row=row, column=25, value=record.temperature_5)
                siaf_sheet.cell(row=row, column=26, value=record.temperature_6)
                
                # Calculate feed per bird
                feed_per_gram_per_bird = 0
                active_male_batches = MaleBirdsStock.objects.filter(status='active')
                active_female_batches = FemaleBirdsStock.objects.filter(status='active')
                
                male_current_birds = sum(batch.get_current_birds() for batch in active_male_batches)
                female_current_birds = sum(batch.get_current_birds() for batch in active_female_batches)
                total_current_birds = male_current_birds + female_current_birds
                
                if total_current_birds > 0:
                    total_feed_kg = (record.feed_morning or 0) + (record.feed_evening or 0)
                    total_feed_grams = total_feed_kg * 1000
                    feed_per_gram_per_bird = round(total_feed_grams / total_current_birds, 2)
                
                # Get mortality counts for the date
                male_mortality_total = MaleBirdsMortality.objects.filter(date=record.date).aggregate(
                    models.Sum('mortality_count'))['mortality_count__sum'] or 0
                female_mortality_total = FemaleBirdsMortality.objects.filter(date=record.date).aggregate(
                    models.Sum('mortality_count'))['mortality_count__sum'] or 0
                total_mortality = male_mortality_total + female_mortality_total
                
                # Feed per bird
                siaf_sheet.cell(row=row, column=27, value=feed_per_gram_per_bird)
                # Male mortality
                siaf_sheet.cell(row=row, column=28, value=male_mortality_total)
                # Female mortality
                siaf_sheet.cell(row=row, column=29, value=female_mortality_total)
                # Total mortality
                siaf_sheet.cell(row=row, column=30, value=total_mortality)
                
                siaf_sheet.cell(row=row, column=31, value=record.medicine)
                siaf_sheet.cell(row=row, column=32, value=record.notes)

            # Add totals for SIAF sheet
            ws = siaf_sheet
            last_row = ws.max_row
            
            # Add a blank row
            last_row += 1
            
            # Add total row with bold formatting
            total_row = last_row + 1
            total_font = Font(bold=True)
            total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            ws.cell(row=total_row, column=1, value="TOTAL").font = total_font
            ws.cell(row=total_row, column=1).fill = total_fill
            
            # Calculate totals for numeric columns
            numeric_cols = [
                2, 3, 4,  # Feed and Water
                5, 6, 7, 8,  # Morning Eggs
                9, 10, 11, 12,  # Evening Eggs
                14, 16, 18, 20,  # AI Hours, Fogger Hours, Fan Hours, Light Hours
                21, 22, 23, 24, 25, 26,  # Temperature readings
                27, 28, 29, 30  # Feed per bird, Male Mortality, Female Mortality, Total Mortality
            ]
            for col in numeric_cols:
                total = 0
                for row in range(2, last_row):  # Start from row 2 to skip header
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)):
                        total += cell_value
                ws.cell(row=total_row, column=col, value=total)
                ws.cell(row=total_row, column=col).font = total_font
                ws.cell(row=total_row, column=col).fill = total_fill

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=siaf_report_{start_date}_to_{end_date}.xlsx'

            # Save the workbook to the response
            wb.save(response)
            return response

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def fetch_record_SIAF(request):
    if request.method == 'GET':
        date_str = request.GET.get('date')
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            record = DailyRecordSIAF.objects.filter(date=date).first()
            
            if record:
                # Calculate feed per bird
                feed_per_gram_per_bird = 0
                total_current_birds = 0
                
                # Get active male and female birds
                active_male_batches = MaleBirdsStock.objects.filter(status='active')
                active_female_batches = FemaleBirdsStock.objects.filter(status='active')
                
                male_current_birds = sum(batch.get_current_birds() for batch in active_male_batches)
                female_current_birds = sum(batch.get_current_birds() for batch in active_female_batches)
                total_current_birds = male_current_birds + female_current_birds
                
                if record and total_current_birds > 0:
                    total_feed_kg = (record.feed_morning or 0) + (record.feed_evening or 0)
                    total_feed_grams = total_feed_kg * 1000
                    feed_per_gram_per_bird = round(total_feed_grams / total_current_birds, 2)
                
                # Get daily closing stock from FeedStock for the date
                # Try to get exact date first, then get the most recent record on or before that date
                feed_stock = FeedStock.objects.filter(date=date).first()
                if not feed_stock:
                    # If no exact match, get the most recent FeedStock on or before this date
                    feed_stock = FeedStock.objects.filter(date__lte=date).order_by('-date').first()
                
                daily_closing_stock = feed_stock.kg if feed_stock else 0
                daily_closing_bundles = feed_stock.bundles if feed_stock else 0
                
                # Get mortality counts for the date
                male_mortality_total = MaleBirdsMortality.objects.filter(date=date).aggregate(
                    models.Sum('mortality_count'))['mortality_count__sum'] or 0
                female_mortality_total = FemaleBirdsMortality.objects.filter(date=date).aggregate(
                    models.Sum('mortality_count'))['mortality_count__sum'] or 0
                total_mortality = male_mortality_total + female_mortality_total
                
                return JsonResponse({
                    'success': True,
                    'data': {
                        'feed_morning': record.feed_morning,
                        'feed_evening': record.feed_evening,
                        'water_intake': record.water_intake,
                        'tray_egg_morning': record.tray_egg_morning,
                        'total_egg_morning': record.total_egg_morning,
                        'damaged_egg_morning': record.damaged_egg_morning,
                        'double_egg_morning': record.double_egg_morning,
                        'tray_egg_evening': record.tray_egg_evening,
                        'total_egg_evening': record.total_egg_evening,
                        'damaged_egg_evening': record.damaged_egg_evening,
                        'double_egg_evening': record.double_egg_evening,
                        'artificial_insemination': record.artificial_insemination,
                        'ai_hours': record.ai_hours,
                        'fogger_used': record.fogger_used,
                        'fogger_hours': record.fogger_hours,
                        'fan_used': record.fan_used,
                        'fan_hours': record.fan_hours,
                        'light_used': record.light_used,
                        'light_hours': record.light_hours,
                        'medicine': record.medicine,
                        'notes': record.notes,
                        'temperature_1': record.temperature_1,
                        'temperature_2': record.temperature_2,
                        'temperature_3': record.temperature_3,
                        'temperature_4': record.temperature_4,
                        'temperature_5': record.temperature_5,
                        'temperature_6': record.temperature_6,
                        'feed_per_gram_per_bird': feed_per_gram_per_bird,
                        'daily_closing_stock': daily_closing_stock,
                        'daily_closing_bundles': daily_closing_bundles,
                        'male_mortality': male_mortality_total,
                        'female_mortality': female_mortality_total,
                        'total_mortality': total_mortality,
                    }
                })
            return JsonResponse({'success': False, 'message': 'No record found for this date'})
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

@login_required
def get_user(request, user_id):
    if request.method == 'GET':
        try:
            user = User.objects.get(id=user_id)
            return JsonResponse({
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'group_id': user.groups.first().id if user.groups.first() else ''
            })
        except User.DoesNotExist:
            return JsonResponse({'error': 'User not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
@login_required()
def add_user(request):
    user_groups = request.user.groups.all()
    u = request.user
    # Get all available groups
    groups = Group.objects.all()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Handle user deletion
        if action == 'delete':
            user_id = request.POST.get('user_id')
            try:
                user = User.objects.get(id=user_id)
                if user.is_superuser:
                    messages.error(request, 'Cannot delete superuser account')
                else:
                    user.delete()
                    messages.success(request, f'User {user.username} deleted successfully')
            except User.DoesNotExist:
                messages.error(request, 'User not found')
            return redirect('add_user')
        
        # Handle user creation/update
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        group_id = request.POST.get('group')

        # Validate passwords match
        if password != confirm_password:
            messages.error(request, 'Passwords do not match')
            return render(request, "add_user.html", {'groups': groups})

        try:
            if user_id:  # Update existing user
                user = User.objects.get(id=user_id)
                # Check if new username conflicts with other users
                if User.objects.filter(username=username).exclude(id=user_id).exists():
                    messages.error(request, 'Username already exists')
                    return redirect('add_user')
                
                # Check if new email conflicts with other users
                if User.objects.filter(email=email).exclude(id=user_id).exists():
                    messages.error(request, 'Email already exists')
                    return redirect('add_user')
                
                user.username = username
                user.email = email
                user.first_name = first_name
                user.last_name = last_name
                if password:  # Only update password if provided
                    user.set_password(password)
                user.save()
                messages.success(request, 'User updated successfully!')
            
            else:  # Create new user
                # Check if username already exists
                if User.objects.filter(username=username).exists():
                    messages.error(request, 'Username already exists')
                    return redirect('add_user')

                # Check if email already exists
                if User.objects.filter(email=email).exists():
                    messages.error(request, 'Email already exists')
                    return redirect('add_user')

                # Create new user
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
            
            # Add user to selected group
            if group_id:
                try:
                    group = Group.objects.get(id=group_id)
                    user.groups.add(group)
                except Group.DoesNotExist:
                    messages.error(request, 'Selected group does not exist.')
                    return render(request, "add_user.html", {'groups': groups})
            
            messages.success(request, 'User created successfully!')
            return redirect('add_user')
        except Exception as e:
            messages.error(request, f'Error creating user: {str(e)}')
            return render(request, "add_user.html", {'groups': groups})

    # For GET request, fetch all users and render the template
    users = User.objects.all().order_by('username')  # Get all users sorted by username
    return render(request, "add_user.html", {
        'groups': groups,
        'user_groups': user_groups,
        'u': u,
        'users': users
    })




# Feed Stock Views
@login_required
def feed_stock_save(request):
    """Save or update feed stock entry"""
    if request.method == 'POST':
        try:
            feed_stock_id = request.POST.get('feed_stock_id')
            date_str = request.POST.get('date')
            kg = float(request.POST.get('kg'))
            notes = request.POST.get('notes', '')

            if not date_str:
                return JsonResponse({'success': False, 'message': 'Date is required'})

            date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Calculate closing stock for this date
            total_stock_received = FeedStock.objects.filter(
                date__lte=date
            ).aggregate(models.Sum('kg'))['kg__sum'] or 0

            # Add the current KG if it's a new entry or if we're updating
            if not feed_stock_id:
                total_stock_received += kg

            # Calculate total feed used up to this date
            siaf_feed_used = DailyRecordSIAF.objects.filter(
                date__lte=date
            ).aggregate(
                morning=models.Sum('feed_morning'),
                evening=models.Sum('feed_evening')
            )
            total_feed_used = (siaf_feed_used['morning'] or 0) + (siaf_feed_used['evening'] or 0)

            if feed_stock_id:
                # Update existing record
                feed_stock = FeedStock.objects.get(id=feed_stock_id)
                feed_stock.date = date
                feed_stock.kg = kg
                feed_stock.notes = notes
                feed_stock.save()
                message = 'Feed stock updated successfully'
            else:
                # Create new record
                feed_stock = FeedStock.objects.create(
                    date=date,
                    kg=kg,
                    notes=notes
                )
                message = 'Feed stock added successfully'

            return JsonResponse({
                'success': True,
                'message': message,
                'entry': {
                    'id': feed_stock.id,
                    'date': feed_stock.date.strftime('%Y-%m-%d'),
                    'kg': feed_stock.kg,
                    'bundles': feed_stock.bundles,
                    'notes': feed_stock.notes
                }
            })
        except FeedStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Feed stock entry not found'})
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_list(request):
    """Get list of recent feed stock entries"""
    if request.method == 'GET':
        try:
            # Get last 10 entries
            entries = FeedStock.objects.all()[:10]
            data = [
                {
                    'id': entry.id,
                    'date': entry.date.strftime('%Y-%m-%d'),
                    'kg': entry.kg,
                    'bundles': entry.bundles,
                    'notes': entry.notes
                }
                for entry in entries
            ]
            return JsonResponse({'success': True, 'entries': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_get(request, feed_stock_id):
    """Get single feed stock entry for editing"""
    if request.method == 'GET':
        try:
            feed_stock = FeedStock.objects.get(id=feed_stock_id)
            return JsonResponse({
                'success': True,
                'entry': {
                    'id': feed_stock.id,
                    'date': feed_stock.date.strftime('%Y-%m-%d'),
                    'kg': feed_stock.kg,
                    'bundles': feed_stock.bundles,
                    'notes': feed_stock.notes
                }
            })
        except FeedStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entry not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_delete(request, feed_stock_id):
    """Delete a feed stock entry"""
    if request.method == 'POST':
        try:
            feed_stock = FeedStock.objects.get(id=feed_stock_id)
            feed_stock.delete()
            return JsonResponse({'success': True, 'message': 'Entry deleted successfully'})
        except FeedStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entry not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_dashboard(request):
    """Get total feed stock and daily closing stock"""
    if request.method == 'GET':
        try:
            selected_date = request.GET.get('date')
            if selected_date:
                selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            else:
                selected_date = timezone.now().date()

            # Calculate total stock received (all FeedStock entries up to selected date)
            total_stock_received = FeedStock.objects.filter(
                date__lte=selected_date
            ).aggregate(models.Sum('kg'))['kg__sum'] or 0

            # Calculate total feed used up to selected date
            siaf_feed_used = DailyRecordSIAF.objects.filter(
                date__lte=selected_date
            ).aggregate(
                morning=models.Sum('feed_morning'),
                evening=models.Sum('feed_evening')
            )
            total_feed_used = (siaf_feed_used['morning'] or 0) + (siaf_feed_used['evening'] or 0)

            # Calculate daily closing stock (total received - total used up to selected date)
            closing_stock_kg = total_stock_received - total_feed_used
            closing_stock_bundles = round(closing_stock_kg / 60, 2)

            # Get today's feed usage
            today_siaf = DailyRecordSIAF.objects.filter(date=selected_date).first()

            today_feed_used = 0
            if today_siaf:
                today_feed_used = (today_siaf.feed_morning or 0) + (today_siaf.feed_evening or 0)

            return JsonResponse({
                'success': True,
                'data': {
                    'date': selected_date.strftime('%Y-%m-%d'),
                    'total_stock_received_kg': round(total_stock_received, 2),
                    'total_stock_received_bundles': round(total_stock_received / 60, 2),
                    'total_feed_used_kg': round(total_feed_used, 2),
                    'total_feed_used_bundles': round(total_feed_used / 60, 2),
                    'closing_stock_kg': round(closing_stock_kg, 2),
                    'closing_stock_bundles': closing_stock_bundles,
                    'today_feed_used_kg': round(today_feed_used, 2),
                    'today_feed_used_bundles': round(today_feed_used / 60, 2)
                }
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_report_data(request):
    """Get feed stock report data for date range"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Missing required parameters'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Query feed stock records for the date range
            records = FeedStock.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('-date')

            # Convert records to list of dictionaries
            data = []
            for record in records:
                data.append({
                    'id': record.id,
                    'date': record.date.strftime('%Y-%m-%d'),
                    'kg': round(record.kg, 2),
                    'bundles': round(record.bundles, 2),
                    'notes': record.notes or ''
                })

            return JsonResponse({
                'success': True,
                'entries': data
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def feed_stock_download_excel(request):
    """Download feed stock report as Excel file"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Missing required parameters'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Query feed stock records
            records = FeedStock.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('date')

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Feed Stock Report"

            # Add headers with styling
            headers = ['Date', 'Weight (KG)', 'Bundles', 'Notes']
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font

            # Add data rows
            total_kg = 0
            total_bundles = 0
            for row_idx, record in enumerate(records, 2):
                ws.cell(row=row_idx, column=1).value = record.date.strftime('%Y-%m-%d')
                ws.cell(row=row_idx, column=2).value = round(record.kg, 2)
                ws.cell(row=row_idx, column=3).value = round(record.bundles, 2)
                ws.cell(row=row_idx, column=4).value = record.notes or ''
                total_kg += record.kg
                total_bundles += record.bundles

            # Add summary section
            summary_row = len(records) + 3
            ws.cell(row=summary_row, column=1).value = "SUMMARY"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)

            summary_data = [
                ('Total KG:', total_kg),
                ('Total Bundles:', total_bundles),
                ('Number of Entries:', len(records)),
                ('Average per Day:', round(total_kg / max(len(set(r.date for r in records)), 1), 2) if records else 0),
                ('Date Range:', f"{start_date} to {end_date}")
            ]

            for idx, (label, value) in enumerate(summary_data, 1):
                ws.cell(row=summary_row + idx, column=1).value = label
                ws.cell(row=summary_row + idx, column=1).font = Font(bold=True)
                ws.cell(row=summary_row + idx, column=2).value = value

            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30

            # Prepare response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Feed_Stock_Report_{start_date}_to_{end_date}.xlsx"'
            wb.save(response)
            return response

        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


# Male Birds Stock Views
@login_required
def male_birds_stock_save(request):
    """Save or update male birds stock entry - supports multiple active batches"""
    if request.method == 'POST':
        try:
            stock_id = request.POST.get('stock_id')
            initial_birds = request.POST.get('initial_birds')
            batch_start_date = request.POST.get('batch_start_date')
            batch_end_date = request.POST.get('batch_end_date')
            notes = request.POST.get('notes', '')

            if not all([initial_birds, batch_start_date]):
                return JsonResponse({'success': False, 'message': 'Initial birds and batch start date are required'})

            initial_birds = int(initial_birds)
            batch_start_date = datetime.strptime(batch_start_date, '%Y-%m-%d').date()
            batch_end_date = datetime.strptime(batch_end_date, '%Y-%m-%d').date() if batch_end_date else None

            if stock_id:
                # Update existing
                stock = MaleBirdsStock.objects.get(id=int(stock_id))
                old_status = stock.status
                
                stock.initial_birds = initial_birds
                stock.batch_start_date = batch_start_date
                stock.batch_end_date = batch_end_date
                stock.notes = notes
                
                # If batch is being ended, calculate and store final mortality
                if batch_end_date and stock.status == 'active':
                    final_mortality = stock.get_current_mortality()
                    stock.final_mortality = final_mortality
                    stock.status = 'ended'
                # If batch end date is removed, revert to active
                elif not batch_end_date and stock.status == 'ended':
                    stock.status = 'active'
                    stock.final_mortality = 0
                
                stock.save()
                return JsonResponse({'success': True, 'message': 'Male birds stock updated successfully', 'id': stock.id})
            else:
                # Create new batch (always starts as active)
                stock = MaleBirdsStock.objects.create(
                    initial_birds=initial_birds,
                    batch_start_date=batch_start_date,
                    batch_end_date=batch_end_date,
                    status='active' if not batch_end_date else 'active',
                    notes=notes
                )
                return JsonResponse({'success': True, 'message': 'Male birds stock added successfully', 'id': stock.id})
        except MaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except ValueError as e:
            return JsonResponse({'success': False, 'message': f'Invalid input: {str(e)}'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_stock_list(request):
    """Get list of male birds stock entries with status and current birds"""
    if request.method == 'GET':
        try:
            stocks = MaleBirdsStock.objects.all().order_by('-batch_start_date')
            data = []
            for stock in stocks:
                current_mortality = stock.get_current_mortality()
                current_birds = stock.get_current_birds()
                
                data.append({
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'current_birds': current_birds,
                    'mortality': current_mortality,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else '',
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else '',
                    'status': stock.status,
                    'notes': stock.notes,
                    'created_at': stock.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_stock_get(request, stock_id):
    """Get single male birds stock entry for editing"""
    if request.method == 'GET':
        try:
            stock = MaleBirdsStock.objects.get(id=stock_id)
            return JsonResponse({
                'success': True,
                'data': {
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else '',
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else '',
                    'notes': stock.notes
                }
            })
        except MaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_stock_delete(request, stock_id):
    """Delete a male birds stock entry"""
    if request.method == 'POST':
        try:
            stock = MaleBirdsStock.objects.get(id=stock_id)
            stock.delete()
            return JsonResponse({'success': True, 'message': 'Male birds stock deleted successfully'})
        except MaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


# Male Birds Mortality Views
@login_required
def male_birds_mortality_save(request):
    """Save or update male birds mortality entry"""
    if request.method == 'POST':
        try:
            mortality_id = request.POST.get('mortality_id')
            date = request.POST.get('date')
            mortality_count = request.POST.get('mortality_count')
            mortality_reason = request.POST.get('mortality_reason', '')

            if not all([date, mortality_count]):
                return JsonResponse({'success': False, 'message': 'Date and mortality count are required'})

            date = datetime.strptime(date, '%Y-%m-%d').date()
            mortality_count = int(mortality_count)

            if mortality_id:
                # Update existing
                mortality = MaleBirdsMortality.objects.get(id=int(mortality_id))
                mortality.date = date
                mortality.mortality_count = mortality_count
                mortality.mortality_reason = mortality_reason
                mortality.save()
                return JsonResponse({'success': True, 'message': 'Mortality record updated successfully', 'id': mortality.id})
            else:
                # Create new
                mortality = MaleBirdsMortality.objects.create(
                    date=date,
                    mortality_count=mortality_count,
                    mortality_reason=mortality_reason
                )
                return JsonResponse({'success': True, 'message': 'Mortality record added successfully', 'id': mortality.id})
        except MaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except ValueError as e:
            return JsonResponse({'success': False, 'message': f'Invalid input: {str(e)}'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_mortality_list(request):
    """Get list of male birds mortality entries"""
    if request.method == 'GET':
        try:
            mortality_records = MaleBirdsMortality.objects.all().order_by('-date')
            data = [{
                'id': record.id,
                'date': record.date.strftime('%Y-%m-%d'),
                'mortality_count': record.mortality_count,
                'mortality_reason': record.mortality_reason,
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            } for record in mortality_records]
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_mortality_get(request, mortality_id):
    """Get single male birds mortality entry for editing"""
    if request.method == 'GET':
        try:
            mortality = MaleBirdsMortality.objects.get(id=mortality_id)
            return JsonResponse({
                'success': True,
                'data': {
                    'id': mortality.id,
                    'date': mortality.date.strftime('%Y-%m-%d'),
                    'mortality_count': mortality.mortality_count,
                    'mortality_reason': mortality.mortality_reason
                }
            })
        except MaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_mortality_delete(request, mortality_id):
    """Delete a male birds mortality entry"""
    if request.method == 'POST':
        try:
            mortality = MaleBirdsMortality.objects.get(id=mortality_id)
            mortality.delete()
            return JsonResponse({'success': True, 'message': 'Mortality record deleted successfully'})
        except MaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_dashboard(request):
    """Get male birds dashboard data - calculates total for all active batches"""
    if request.method == 'GET':
        try:
            # Get all active batches
            active_stocks = MaleBirdsStock.objects.filter(status='active').order_by('-batch_start_date')
            
            if not active_stocks.exists():
                return JsonResponse({
                    'success': True,
                    'data': {
                        'total_initial_birds': 0,
                        'total_current_birds': 0,
                        'total_mortality': 0,
                        'active_batches': 0,
                        'ended_batches': 0,
                        'batches': []
                    }
                })
            
            today = timezone.now().date()
            total_initial_birds = 0
            total_current_birds = 0
            total_mortality = 0
            batches_detail = []
            
            # Calculate totals for all active batches
            for stock in active_stocks:
                current_mortality = stock.get_current_mortality()
                current_birds = stock.get_current_birds()
                days_running = (today - stock.batch_start_date).days if stock.batch_start_date else 0
                
                total_initial_birds += stock.initial_birds
                total_current_birds += current_birds
                total_mortality += current_mortality
                
                batches_detail.append({
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'current_birds': current_birds,
                    'mortality': current_mortality,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else None,
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else None,
                    'days_running': days_running,
                    'status': stock.status
                })
            
            # Count active and ended batches
            active_count = MaleBirdsStock.objects.filter(status='active').count()
            ended_count = MaleBirdsStock.objects.filter(status='ended').count()
            
            return JsonResponse({
                'success': True,
                'data': {
                    'total_initial_birds': total_initial_birds,
                    'total_current_birds': total_current_birds,
                    'total_mortality': total_mortality,
                    'active_batches': active_count,
                    'ended_batches': ended_count,
                    'batches': batches_detail
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_report_data(request):
    """Get male birds report data with date filtering"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Start date and end date are required'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            mortality_records = MaleBirdsMortality.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('-date')

            data = [{
                'date': record.date.strftime('%Y-%m-%d'),
                'mortality_count': record.mortality_count,
                'mortality_reason': record.mortality_reason
            } for record in mortality_records]

            return JsonResponse({
                'success': True,
                'records': data
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def male_birds_download_excel(request):
    """Download male birds report as Excel file with multiple sheets"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Start date and end date are required'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # ===== SHEET 1: Mortality Report =====
            ws_mortality = wb.create_sheet('Mortality Report')
            
            # Get mortality records for the date range
            mortality_records = MaleBirdsMortality.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('date')

            # Add headers
            headers = ['Date', 'Mortality Count', 'Mortality Reason']
            ws_mortality.append(headers)

            # Style headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws_mortality[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add data
            cumulative_mortality = 0
            for record in mortality_records:
                cumulative_mortality += record.mortality_count
                ws_mortality.append([
                    record.date.strftime('%Y-%m-%d'),
                    record.mortality_count,
                    record.mortality_reason or ''
                ])

            # Add summary section
            ws_mortality.append([])
            ws_mortality.append(['Summary'])
            ws_mortality.append(['Total Mortality:', cumulative_mortality])

            # Adjust column widths
            ws_mortality.column_dimensions['A'].width = 15
            ws_mortality.column_dimensions['B'].width = 15
            ws_mortality.column_dimensions['C'].width = 30

            # ===== SHEET 2: Batch History =====
            ws_batch = wb.create_sheet('Batch History')

            # Get all male birds batches
            all_batches = MaleBirdsStock.objects.all().order_by('-batch_start_date')

            # Add headers
            batch_headers = ['Start Date', 'End Date', 'Initial Birds', 'Status', 'Total Mortality', 'Current Birds', 'Notes']
            ws_batch.append(batch_headers)

            # Style headers
            for cell in ws_batch[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add batch data
            for batch in all_batches:
                batch_mortality = batch.get_current_mortality()
                current_birds = batch.get_current_birds()
                ws_batch.append([
                    batch.batch_start_date.strftime('%Y-%m-%d') if batch.batch_start_date else '',
                    batch.batch_end_date.strftime('%Y-%m-%d') if batch.batch_end_date else '',
                    batch.initial_birds,
                    batch.status,
                    batch_mortality,
                    current_birds,
                    batch.notes or ''
                ])

            # Adjust column widths for batch sheet
            ws_batch.column_dimensions['A'].width = 15
            ws_batch.column_dimensions['B'].width = 15
            ws_batch.column_dimensions['C'].width = 15
            ws_batch.column_dimensions['D'].width = 12
            ws_batch.column_dimensions['E'].width = 15
            ws_batch.column_dimensions['F'].width = 15
            ws_batch.column_dimensions['G'].width = 25

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="male_birds_report_{start_date}_{end_date}.xlsx"'
            wb.save(response)
            return response

        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


# Female Birds Stock Views
@login_required
def female_birds_stock_save(request):
    """Save or update female birds stock entry - supports multiple active batches"""
    if request.method == 'POST':
        try:
            stock_id = request.POST.get('stock_id')
            initial_birds = request.POST.get('initial_birds')
            batch_start_date = request.POST.get('batch_start_date')
            batch_end_date = request.POST.get('batch_end_date')
            notes = request.POST.get('notes', '')

            if not all([initial_birds, batch_start_date]):
                return JsonResponse({'success': False, 'message': 'Initial birds and batch start date are required'})

            initial_birds = int(initial_birds)
            batch_start_date = datetime.strptime(batch_start_date, '%Y-%m-%d').date()
            batch_end_date = datetime.strptime(batch_end_date, '%Y-%m-%d').date() if batch_end_date else None

            if stock_id:
                # Update existing
                stock = FemaleBirdsStock.objects.get(id=int(stock_id))
                old_status = stock.status
                
                stock.initial_birds = initial_birds
                stock.batch_start_date = batch_start_date
                stock.batch_end_date = batch_end_date
                stock.notes = notes
                
                # If batch is being ended, calculate and store final mortality
                if batch_end_date and stock.status == 'active':
                    final_mortality = stock.get_current_mortality()
                    stock.final_mortality = final_mortality
                    stock.status = 'ended'
                # If batch end date is removed, revert to active
                elif not batch_end_date and stock.status == 'ended':
                    stock.status = 'active'
                    stock.final_mortality = 0
                
                stock.save()
                return JsonResponse({'success': True, 'message': 'Female birds stock updated successfully', 'id': stock.id})
            else:
                # Create new batch (always starts as active)
                stock = FemaleBirdsStock.objects.create(
                    initial_birds=initial_birds,
                    batch_start_date=batch_start_date,
                    batch_end_date=batch_end_date,
                    status='active' if not batch_end_date else 'active',
                    notes=notes
                )
                return JsonResponse({'success': True, 'message': 'Female birds stock added successfully', 'id': stock.id})
        except FemaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except ValueError as e:
            return JsonResponse({'success': False, 'message': f'Invalid input: {str(e)}'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_stock_list(request):
    """Get list of female birds stock entries with status and current birds"""
    if request.method == 'GET':
        try:
            stocks = FemaleBirdsStock.objects.all().order_by('-batch_start_date')
            data = []
            for stock in stocks:
                current_mortality = stock.get_current_mortality()
                current_birds = stock.get_current_birds()
                
                data.append({
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'current_birds': current_birds,
                    'mortality': current_mortality,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else '',
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else '',
                    'status': stock.status,
                    'notes': stock.notes,
                    'created_at': stock.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_stock_get(request, stock_id):
    """Get single female birds stock entry for editing"""
    if request.method == 'GET':
        try:
            stock = FemaleBirdsStock.objects.get(id=stock_id)
            return JsonResponse({
                'success': True,
                'data': {
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else '',
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else '',
                    'notes': stock.notes
                }
            })
        except FemaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_stock_delete(request, stock_id):
    """Delete a female birds stock entry"""
    if request.method == 'POST':
        try:
            stock = FemaleBirdsStock.objects.get(id=stock_id)
            stock.delete()
            return JsonResponse({'success': True, 'message': 'Female birds stock deleted successfully'})
        except FemaleBirdsStock.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Stock record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


# Female Birds Mortality Views
@login_required
def female_birds_mortality_save(request):
    """Save or update female birds mortality entry"""
    if request.method == 'POST':
        try:
            mortality_id = request.POST.get('mortality_id')
            date = request.POST.get('date')
            mortality_count = request.POST.get('mortality_count')
            mortality_reason = request.POST.get('mortality_reason', '')

            if not all([date, mortality_count]):
                return JsonResponse({'success': False, 'message': 'Date and mortality count are required'})

            date = datetime.strptime(date, '%Y-%m-%d').date()
            mortality_count = int(mortality_count)

            if mortality_id:
                # Update existing
                mortality = FemaleBirdsMortality.objects.get(id=int(mortality_id))
                mortality.date = date
                mortality.mortality_count = mortality_count
                mortality.mortality_reason = mortality_reason
                mortality.save()
                return JsonResponse({'success': True, 'message': 'Mortality record updated successfully', 'id': mortality.id})
            else:
                # Create new
                mortality = FemaleBirdsMortality.objects.create(
                    date=date,
                    mortality_count=mortality_count,
                    mortality_reason=mortality_reason
                )
                return JsonResponse({'success': True, 'message': 'Mortality record added successfully', 'id': mortality.id})
        except FemaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except ValueError as e:
            return JsonResponse({'success': False, 'message': f'Invalid input: {str(e)}'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_mortality_list(request):
    """Get list of female birds mortality entries"""
    if request.method == 'GET':
        try:
            mortality_records = FemaleBirdsMortality.objects.all().order_by('-date')
            data = [{
                'id': record.id,
                'date': record.date.strftime('%Y-%m-%d'),
                'mortality_count': record.mortality_count,
                'mortality_reason': record.mortality_reason,
                'created_at': record.created_at.strftime('%Y-%m-%d %H:%M:%S')
            } for record in mortality_records]
            return JsonResponse({'success': True, 'data': data})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_mortality_get(request, mortality_id):
    """Get single female birds mortality entry for editing"""
    if request.method == 'GET':
        try:
            mortality = FemaleBirdsMortality.objects.get(id=mortality_id)
            return JsonResponse({
                'success': True,
                'data': {
                    'id': mortality.id,
                    'date': mortality.date.strftime('%Y-%m-%d'),
                    'mortality_count': mortality.mortality_count,
                    'mortality_reason': mortality.mortality_reason
                }
            })
        except FemaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_mortality_delete(request, mortality_id):
    """Delete a female birds mortality entry"""
    if request.method == 'POST':
        try:
            mortality = FemaleBirdsMortality.objects.get(id=mortality_id)
            mortality.delete()
            return JsonResponse({'success': True, 'message': 'Mortality record deleted successfully'})
        except FemaleBirdsMortality.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Mortality record not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_dashboard(request):
    """Get female birds dashboard data - calculates total for all active batches"""
    if request.method == 'GET':
        try:
            # Get all active batches
            active_stocks = FemaleBirdsStock.objects.filter(status='active').order_by('-batch_start_date')
            
            if not active_stocks.exists():
                return JsonResponse({
                    'success': True,
                    'data': {
                        'total_initial_birds': 0,
                        'total_current_birds': 0,
                        'total_mortality': 0,
                        'active_batches': 0,
                        'ended_batches': 0,
                        'batches': []
                    }
                })
            
            today = timezone.now().date()
            total_initial_birds = 0
            total_current_birds = 0
            total_mortality = 0
            batches_detail = []
            
            # Calculate totals for all active batches
            for stock in active_stocks:
                current_mortality = stock.get_current_mortality()
                current_birds = stock.get_current_birds()
                days_running = (today - stock.batch_start_date).days if stock.batch_start_date else 0
                
                total_initial_birds += stock.initial_birds
                total_current_birds += current_birds
                total_mortality += current_mortality
                
                batches_detail.append({
                    'id': stock.id,
                    'initial_birds': stock.initial_birds,
                    'current_birds': current_birds,
                    'mortality': current_mortality,
                    'batch_start_date': stock.batch_start_date.strftime('%Y-%m-%d') if stock.batch_start_date else None,
                    'batch_end_date': stock.batch_end_date.strftime('%Y-%m-%d') if stock.batch_end_date else None,
                    'days_running': days_running,
                    'status': stock.status
                })
            
            # Count active and ended batches
            active_count = FemaleBirdsStock.objects.filter(status='active').count()
            ended_count = FemaleBirdsStock.objects.filter(status='ended').count()
            
            return JsonResponse({
                'success': True,
                'data': {
                    'total_initial_birds': total_initial_birds,
                    'total_current_birds': total_current_birds,
                    'total_mortality': total_mortality,
                    'active_batches': active_count,
                    'ended_batches': ended_count,
                    'batches': batches_detail
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_report_data(request):
    """Get female birds report data with date filtering"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Start date and end date are required'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            mortality_records = FemaleBirdsMortality.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('-date')

            data = [{
                'date': record.date.strftime('%Y-%m-%d'),
                'mortality_count': record.mortality_count,
                'mortality_reason': record.mortality_reason
            } for record in mortality_records]

            return JsonResponse({
                'success': True,
                'records': data
            })
        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def female_birds_download_excel(request):
    """Download female birds report as Excel file with multiple sheets"""
    if request.method == 'GET':
        try:
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')

            if not all([start_date, end_date]):
                return JsonResponse({'success': False, 'message': 'Start date and end date are required'})

            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # ===== SHEET 1: Mortality Report =====
            ws_mortality = wb.create_sheet('Mortality Report')
            
            # Get mortality records for the date range
            mortality_records = FemaleBirdsMortality.objects.filter(
                date__range=[start_date, end_date]
            ).order_by('date')

            # Add headers
            headers = ['Date', 'Mortality Count', 'Mortality Reason']
            ws_mortality.append(headers)

            # Style headers
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws_mortality[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add data
            cumulative_mortality = 0
            for record in mortality_records:
                cumulative_mortality += record.mortality_count
                ws_mortality.append([
                    record.date.strftime('%Y-%m-%d'),
                    record.mortality_count,
                    record.mortality_reason or ''
                ])

            # Add summary section
            ws_mortality.append([])
            ws_mortality.append(['Summary'])
            ws_mortality.append(['Total Mortality:', cumulative_mortality])

            # Adjust column widths
            ws_mortality.column_dimensions['A'].width = 15
            ws_mortality.column_dimensions['B'].width = 15
            ws_mortality.column_dimensions['C'].width = 30

            # ===== SHEET 2: Batch History =====
            ws_batch = wb.create_sheet('Batch History')

            # Get all female birds batches
            all_batches = FemaleBirdsStock.objects.all().order_by('-batch_start_date')

            # Add headers
            batch_headers = ['Start Date', 'End Date', 'Initial Birds', 'Status', 'Total Mortality', 'Current Birds', 'Notes']
            ws_batch.append(batch_headers)

            # Style headers
            for cell in ws_batch[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Add batch data
            for batch in all_batches:
                batch_mortality = batch.get_current_mortality()
                current_birds = batch.get_current_birds()
                ws_batch.append([
                    batch.batch_start_date.strftime('%Y-%m-%d') if batch.batch_start_date else '',
                    batch.batch_end_date.strftime('%Y-%m-%d') if batch.batch_end_date else '',
                    batch.initial_birds,
                    batch.status,
                    batch_mortality,
                    current_birds,
                    batch.notes or ''
                ])

            # Adjust column widths for batch sheet
            ws_batch.column_dimensions['A'].width = 15
            ws_batch.column_dimensions['B'].width = 15
            ws_batch.column_dimensions['C'].width = 15
            ws_batch.column_dimensions['D'].width = 12
            ws_batch.column_dimensions['E'].width = 15
            ws_batch.column_dimensions['F'].width = 15
            ws_batch.column_dimensions['G'].width = 25

            # Create response
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="female_birds_report_{start_date}_{end_date}.xlsx"'
            wb.save(response)
            return response

        except ValueError:
            return JsonResponse({'success': False, 'message': 'Invalid date format'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return JsonResponse({'success': False, 'message': 'Invalid request method'})
